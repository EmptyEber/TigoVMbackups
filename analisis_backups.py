import os
import logging
from collections import defaultdict, Counter
from datetime import datetime
from dateutil import parser
from PIL import Image, ImageTk, ImageDraw, ImageFont
import openpyxl
from openpyxl.styles import PatternFill
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import sys

# --- Función para gestionar rutas de recursos (Esencial para PyInstaller) ---
def obtener_ruta_recursos(relative_path):
    """
    Obtiene la ruta absoluta a un recurso, ya sea en el entorno de desarrollo
    o dentro del paquete PyInstaller.
    """
    if hasattr(sys, '_MEIPASS'): # Si el script está empaquetado por PyInstaller
        # _MEIPASS es la ruta temporal donde PyInstaller extrae los archivos
        return os.path.join(sys._MEIPASS, relative_path)
    # Si se ejecuta como script normal, usa la ruta relativa al directorio del script
    return os.path.join(os.path.dirname(__file__), relative_path)

# --- CONFIGURACIÓN ---
RUTA_INFORMES = "InformesSinProcesar"
RUTA_EXPORTACION = "InformesExportados"
CONFIGURACION_FECHA_HORA = '%Y-%m-%d %H:%M:%S'

# --- Logging (para registro de eventos y errores) ---
logging.basicConfig(filename='log_analisis_backups_app.txt', level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')


def buscar_encabezados(sheet):
    # Definición de encabezados esperados y sus posibles aliases
    posibles = {
        'object name': ['object name', 'servidor', 'server', 'name', 'hostname'], # Añadido 'hostname'
        'job name': ['job name', 'nombre del job', 'job', 'jobname'], # Añadido 'jobname'
        'start time': ['start time', 'hora de inicio', 'start'],
        'finish time': ['finish time', 'hora de fin', 'finish'],
        'duration': ['duration', 'duración'],
        'data read, gb': ['data read, gb', 'datos leídos', 'data read'],
        'actual total backup size, gb': ['actual total backup size, gb', 'tamaño backup', 'backup size'],
        'backup status': ['backup status', 'estado', 'status']
    }

    # Búsqueda en las primeras 20 filas para encontrar la fila de encabezados
    for row_idx in range(1, 21):
        row = sheet[row_idx]
        encabezados_encontrados = {}
        for col_idx, cell in enumerate(row):
            valor = str(cell.value).strip().lower() if cell.value else ""
            for clave_esperada, aliases in posibles.items():
                # Importante: solo añadir si aún no se ha encontrado esa clave esperada
                if valor in aliases and clave_esperada not in encabezados_encontrados:
                    encabezados_encontrados[clave_esperada] = col_idx # Almacena el índice de la columna
                    break # Una vez que encontramos un alias para esta celda, pasamos a la siguiente celda

        # Si se encontraron todos los encabezados requeridos, se devuelve el diccionario y la fila de inicio de datos
        if len(encabezados_encontrados) == len(posibles):
            return encabezados_encontrados, row_idx + 1 # row_idx + 1 es la siguiente fila después de los encabezados

    return None, None # Si no se encuentran todos los encabezados en las primeras 20 filas


def filtrar_fallos_reales(backups_por_servidor):
    filtrados = defaultdict(list)
    agrupados = defaultdict(list)
    for servidor, historial in backups_por_servidor.items():
        for b in historial:
            clave = (servidor, b['nombre_trabajo'], b['inicio'].date())
            agrupados[clave].append(b)

    for clave, ejecuciones in agrupados.items():
        # Si NO hay NINGUNA ejecución exitosa para esa clave (servidor, trabajo, día),
        # entonces consideramos todas las ejecuciones de esa clave como una "falla real"
        if not any(e['estado'] == 'success' for e in ejecuciones):
            for e in ejecuciones:
                filtrados[clave[0]].append(e)
    return filtrados


def analizar_informes(ruta_informes):
    backups_por_servidor = defaultdict(list)
    trabajos_por_servidor = defaultdict(set)
    fechas_por_servidor_y_trabajo = defaultdict(set)
    trabajos = set()
    servidores = set()
    fechas = set()

    archivos = [os.path.join(ruta_informes, f) for f in os.listdir(ruta_informes) if f.endswith('.xlsx')]
    if not archivos:
        logging.warning(f"No hay archivos XLSX en la carpeta de informes: {ruta_informes}")
        return backups_por_servidor, trabajos_por_servidor, trabajos, servidores, fechas, fechas_por_servidor_y_trabajo

    for archivo in archivos:
        try:
            workbook = openpyxl.load_workbook(archivo, data_only=True)
            sheet = workbook.active
            headers, fila_inicio = buscar_encabezados(sheet)
            if not headers:
                logging.error(f"No se encontraron todos los encabezados válidos en {archivo}. Se saltará este archivo.")
                continue

            for row_idx, row_values in enumerate(sheet.iter_rows(min_row=fila_inicio, values_only=True)):
                try:
                    # Acceso a los valores usando los índices de columna encontrados
                    servidor = str(row_values[headers['object name']]).strip()
                    if not servidor: # Saltar filas sin nombre de servidor
                        continue

                    # Asegurar que las fechas se parseen correctamente, manejando posibles tipos de datos de Excel
                    start_time_val = row_values[headers['start time']]
                    finish_time_val = row_values[headers['finish time']]

                    fecha_inicio = parser.parse(str(start_time_val)) if isinstance(start_time_val, (str, datetime)) else None
                    fecha_fin = parser.parse(str(finish_time_val)) if isinstance(finish_time_val, (str, datetime)) else None

                    if not fecha_inicio or not fecha_fin: # Saltar si las fechas no son válidas
                        logging.warning(f"Fila {fila_inicio + row_idx} en {archivo}: Fecha de inicio/fin inválida ('{start_time_val}'/'{finish_time_val}'). Se saltará.")
                        continue

                    estado = str(row_values[headers['backup status']]).lower().strip()
                    trabajo = str(row_values[headers['job name']]).strip()

                    backup = {
                        'nombre_trabajo': trabajo,
                        'inicio': fecha_inicio,
                        'fin': fecha_fin,
                        'duracion': row_values[headers['duration']],
                        'data_read': row_values[headers['data read, gb']],
                        'tamano_backup': row_values[headers['actual total backup size, gb']],
                        'estado': estado
                    }

                    backups_por_servidor[servidor].append(backup)
                    trabajos.add(trabajo)
                    trabajos_por_servidor[servidor].add(trabajo)
                    servidores.add(servidor)
                    fechas.add(fecha_inicio.date())
                    fechas_por_servidor_y_trabajo[(servidor, trabajo)].add(fecha_inicio.date())
                except Exception as e:
                    logging.warning(f"Error procesando fila {fila_inicio + row_idx} de {archivo}: {e}. Datos de fila: {row_values}")
        except Exception as e:
            logging.error(f"No se pudo procesar el archivo {archivo}: {e}")

    backups_filtrados = filtrar_fallos_reales(backups_por_servidor)
    return backups_filtrados, trabajos_por_servidor, sorted(trabajos), sorted(servidores), sorted(fechas), fechas_por_servidor_y_trabajo


def exportar_excel(resultados):
    if not os.path.exists(RUTA_EXPORTACION):
        os.makedirs(RUTA_EXPORTACION)

    fecha_actual = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    ruta_archivo = os.path.join(RUTA_EXPORTACION, f"InformeGenerado_{fecha_actual}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoría Fallas"
    ws.append(["Servidor", "Nombre Trabajo", "Inicio", "Fin", "Tamaño Backup (GB)", "Estado"])

    colores = {
        "success": "C6EFCE", # Mantener por consistencia, aunque no se esperan aquí
        "failed": "FFC7CE",
        "warning": "FFEB9C"
    }

    resumen_estado = Counter()
    fallas_por_dia_servidor_trabajo = {}
    for r in resultados:
        clave_diaria = (r['servidor'], r['nombre_trabajo'], r['inicio'].date())
        # Guarda solo la última ocurrencia del día si hay varias, o la primera si es la única
        if clave_diaria not in fallas_por_dia_servidor_trabajo or r['inicio'] > fallas_por_dia_servidor_trabajo[clave_diaria]['inicio']:
            fallas_por_dia_servidor_trabajo[clave_diaria] = r

    for r in fallas_por_dia_servidor_trabajo.values():
        fila = [
            r['servidor'], r['nombre_trabajo'],
            r['inicio'].strftime(CONFIGURACION_FECHA_HORA),
            r['fin'].strftime(CONFIGURACION_FECHA_HORA),
            r['tamano_backup'], r['estado']
        ]
        ws.append(fila)
        resumen_estado[r['estado']] += 1
        color = colores.get(r['estado'], "FFFFFF")
        for col in range(1, len(fila) + 1):
            ws.cell(row=ws.max_row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    ws_resumen = wb.create_sheet(title="Resumen Fallas")
    ws_resumen.append(["Estado", "Cantidad"])
    for estado, cantidad in resumen_estado.items():
        ws_resumen.append([estado, cantidad])

    wb.save(ruta_archivo)
    print(f"Informe de Fallas exportado en: {ruta_archivo}")


def crear_interfaz(backups_por_servidor, trabajos_por_servidor, trabajos, servidores, fechas, fechas_por_servidor_y_trabajo):
    app = ttk.Window(themename="superhero")
    app.title("Explorador de Backups")
    app.geometry("1300x750")

    # MODIFICACIÓN CLAVE AQUÍ: Usar PNG para el icono de la aplicación
    # Asegúrate de tener un archivo ServidorICONO_app.png en la misma carpeta que tu script
    try:
        icono_app_png = Image.open(obtener_ruta_recursos("ServidorICONO_app.png"))
        icono_app_png = Image.open(obtener_ruta_recursos("ServidorICONO_app.png"))
        # Es buena idea redimensionar para asegurar que el PNG sea un tamaño razonable para un icono
        icono_app_png = icono_app_png.resize((64, 64), Image.LANCZOS) # Ejemplo: 64x64, o el tamaño que prefieras
        icono_tk_photo = ImageTk.PhotoImage(icono_app_png)
        # Establece el icono de la ventana principal y la barra de tareas
        app.iconphoto(True, icono_tk_photo)
    except Exception as e:
        # En caso de que el PNG falle, intentamos cargar el ICO original (que ha dado problemas)
        # O simplemente no ponemos icono si este también falla.
        print(f"ATENCIÓN: Fallo al cargar ServidorICONO_app.png como icono de la aplicación: {e}")
        try:
            # Si el PNG falla, intentamos usar el ICO original para la barra de título
            app.iconbitmap(obtener_ruta_recursos("ServidorICONO.ico"))
        except Exception as e_ico:
            print(f"ATENCIÓN: Fallo al cargar ServidorICONO.ico como icono de barra de título: {e_ico}")


    # --- Frame superior con ícono y título ---
    frame_top = ttk.Frame(app)
    frame_top.pack(fill="x", padx=10, pady=10)

    try:
        icono_buscar = Image.open(obtener_ruta_recursos("ICONObuscar.png"))
        icono_buscar = icono_buscar.resize((30, 30), Image.LANCZOS)
        icono_tk = ImageTk.PhotoImage(icono_buscar)
        label_icono = ttk.Label(frame_top, image=icono_tk)
        label_icono.image = icono_tk
        label_icono.pack(side="left", padx=(5, 10))
    except Exception as e:
        print(f"ATENCIÓN: No se pudo cargar ICONObuscar.png: {e}")

    label_titulo = ttk.Label(frame_top, text="Explorador de Backups", font=("Segoe UI", 18, "bold"))
    label_titulo.pack(side="left")

    # --- Frame de filtros ---
    frame_filtros = ttk.LabelFrame(app, text="Filtros de Búsqueda", padding=15)
    frame_filtros.pack(padx=10, pady=5, fill="x")

    # Filtros de búsqueda
    ttk.Label(frame_filtros, text="Servidor").grid(row=1, column=0, sticky="w")
    combo_servidor = ttk.Combobox(frame_filtros, values=servidores, state="readonly", width=25)
    combo_servidor.grid(row=1, column=1, padx=5)

    ttk.Label(frame_filtros, text="Trabajo").grid(row=1, column=2, sticky="w")
    combo_trabajo = ttk.Combobox(frame_filtros, values=[], state="readonly", width=25)
    combo_trabajo.grid(row=1, column=3, padx=5)

    ttk.Label(frame_filtros, text="Fecha").grid(row=1, column=4, sticky="w")
    combo_fecha = ttk.Combobox(frame_filtros, values=[], state="readonly", width=20)
    combo_fecha.grid(row=1, column=5, padx=5)

    ttk.Label(frame_filtros, text="Estado").grid(row=1, column=6, sticky="w")
    combo_estado = ttk.Combobox(frame_filtros, values=["", "success", "failed", "warning"], state="readonly", width=15)
    combo_estado.grid(row=1, column=7, padx=5)

    # Botones
    ttk.Button(frame_filtros, text="Limpiar Filtros", command=lambda: limpiar_filtros(), bootstyle=INFO).grid(row=2, column=1, pady=10)
    ttk.Button(frame_filtros, text="Buscar", command=lambda: buscar(), bootstyle=PRIMARY).grid(row=2, column=3, pady=10)
    ttk.Button(frame_filtros, text="Exportar a Excel", command=lambda: exportar(), bootstyle=SUCCESS).grid(row=2, column=4, pady=10)

    # Logo SAVIA con créditos, mini y más a la derecha
    try:
        ruta_logo = obtener_ruta_recursos("logoSAVIA.png")
        logo_img = Image.open(ruta_logo).convert("RGBA")
        ancho, alto = logo_img.size
        nuevo_alto = alto + 25
        lienzo = Image.new("RGBA", (ancho, nuevo_alto), (255, 255, 255, 0))
        lienzo.paste(logo_img, (0, 0))

        draw = ImageDraw.Draw(lienzo)
        texto = "Desarrollado por Eber Romero"
        fuente = ImageFont.load_default() # O puedes especificar una fuente como ImageFont.truetype("arial.ttf", 10)
        x = (ancho - draw.textlength(texto, font=fuente)) // 2
        y = alto + 5
        draw.text((x, y), texto, fill="white", font=fuente)

        lienzo = lienzo.resize((120, int(nuevo_alto * 120 / ancho)), Image.LANCZOS)
        logo_tk = ImageTk.PhotoImage(lienzo)

        label_logo = ttk.Label(frame_filtros, image=logo_tk)
        label_logo.image = logo_tk
        label_logo.grid(row=0, column=10, rowspan=3, padx=30, pady=10, sticky="ns")
    except Exception as e:
        print(f"ATENCIÓN: No se pudo cargar el logo SAVIA: {e}")

    # Área de resultados
    text_resultado = ttk.ScrolledText(app, width=150, height=30, wrap="none", font=("Consolas", 10))
    text_resultado.pack(padx=10, pady=10, fill="both", expand=True)

    # --- Funciones auxiliares para la UI (actualizan los ComboBoxes) ---
    def actualizar_combobox_trabajos(event=None):
        servidor_seleccionado = combo_servidor.get()
        if servidor_seleccionado and servidor_seleccionado in trabajos_por_servidor:
            jobs_filtrados = sorted(list(trabajos_por_servidor[servidor_seleccionado]))
            combo_trabajo['values'] = [""] + jobs_filtrados # Añadir opción "Todos"
            combo_trabajo.set("") # Limpiar selección anterior
            actualizar_combobox_fechas() # Actualizar fechas al cambiar el servidor
        else:
            # Si no hay servidor seleccionado o no se encuentra, mostrar todos los trabajos
            combo_trabajo['values'] = [""] + sorted(list(trabajos)) # Todos los trabajos de todos los servidores
            combo_trabajo.set("")
            combo_fecha['values'] = [""] + sorted([f.strftime('%Y-%m-%d') for f in fechas]) # Todas las fechas
            combo_fecha.set("")

    def actualizar_combobox_fechas(event=None):
        servidor_seleccionado = combo_servidor.get()
        trabajo_seleccionado = combo_trabajo.get()
        fechas_filtradas = set()

        if servidor_seleccionado and trabajo_seleccionado:
            clave = (servidor_seleccionado, trabajo_seleccionado)
            if clave in fechas_por_servidor_y_trabajo:
                fechas_filtradas.update(fechas_por_servidor_y_trabajo[clave])
        elif servidor_seleccionado:
            for job in trabajos_por_servidor[servidor_seleccionado]:
                clave = (servidor_seleccionado, job)
                if clave in fechas_por_servidor_y_trabajo:
                    fechas_filtradas.update(fechas_por_servidor_y_trabajo[clave])
        else:
            # Si no hay servidor ni trabajo seleccionado, muestra todas las fechas globales
            fechas_filtradas.update(fechas) # 'fechas' contiene todas las fechas cargadas

        # Formatear fechas a string y ordenar
        fechas_formateadas = sorted([f.strftime('%Y-%m-%d') for f in fechas_filtradas])
        combo_fecha['values'] = [""] + fechas_formateadas
        combo_fecha.set("")


    # Conectar comboboxes para actualizar las opciones dinámicamente
    combo_servidor.bind("<<ComboboxSelected>>", actualizar_combobox_trabajos)
    combo_trabajo.bind("<<ComboboxSelected>>", actualizar_combobox_fechas)

    # Inicializar comboboxes al inicio con todos los datos disponibles
    combo_servidor['values'] = [""] + sorted(list(servidores)) # Llenar el combo de servidores
    actualizar_combobox_trabajos() # Llama para que se llenen los otros combos inicialmente
    actualizar_combobox_fechas()


    def limpiar_filtros():
        combo_servidor.set("")
        combo_trabajo.set("")
        combo_fecha.set("")
        combo_estado.set("")
        text_resultado.delete(1.0, ttk.END)
        # Re-actualizar comboboxes para mostrar todas las opciones
        actualizar_combobox_trabajos()
        actualizar_combobox_fechas()


    def buscar():
        text_resultado.delete(1.0, ttk.END)
        filtro_srv = combo_servidor.get().strip()
        filtro_job = combo_trabajo.get().strip()
        filtro_fecha = combo_fecha.get().strip()
        filtro_estado = combo_estado.get().strip()

        resultados_filtrados = []
        resumen_diario_fallas = {}

        # Iterar sobre los backups_por_servidor que ya están pre-filtrados por fallas reales
        for servidor, historial in backups_por_servidor.items():
            if filtro_srv and filtro_srv != servidor:
                continue
            
            for intento in historial:
                # La lógica de 'estado == success' ya se manejó en filtrar_fallos_reales
                # Aquí solo aplicamos los filtros seleccionados en la UI
                if filtro_job and filtro_job != intento['nombre_trabajo']:
                    continue
                if filtro_fecha and filtro_fecha != intento['inicio'].strftime('%Y-%m-%d'):
                    continue
                # Se eliminó el 'continue' incorrecto que estaba aquí, permitiendo que la verificación de estado ocurra.
                if filtro_estado and filtro_estado != intento['estado']:
                    continue

                resultados_filtrados.append({**intento, 'servidor': servidor})

                clave_diaria = (servidor, intento['nombre_trabajo'], intento['inicio'].date())
                # Solo se actualiza el resumen si el intento actual es más reciente para esa clave diaria
                if clave_diaria not in resumen_diario_fallas or intento['inicio'] > resumen_diario_fallas[clave_diaria]['inicio']:
                    resumen_diario_fallas[clave_diaria] = {**intento, 'servidor': servidor}


        if resultados_filtrados:
            resultados_filtrados.sort(key=lambda x: (x['servidor'], x['nombre_trabajo'], x['inicio']))

            text_resultado.insert(ttk.END, "Todos los intentos (filtrados por falla):\n")
            for r in resultados_filtrados:
                text_resultado.insert(ttk.END, f"{r['inicio'].strftime('%Y-%m-%d %H:%M:%S')} | {r['servidor']} | {r['nombre_trabajo']} | Estado: {r['estado']}\n")

            text_resultado.insert(ttk.END, "\n--- Resumen de Fallas Diarias (por Servidor y Trabajo) ---\n")
            fallas_para_mostrar = sorted(resumen_diario_fallas.values(), key=lambda x: (x['servidor'], x['nombre_trabajo'], x['inicio'].date()))

            if fallas_para_mostrar:
                for r in fallas_para_mostrar:
                    text_resultado.insert(ttk.END, f" {r['inicio'].strftime('%Y-%m-%d')} | Servidor: {r['servidor']} | Trabajo: {r['nombre_trabajo']} | Estado: {r['estado']}\n")
            else:
                text_resultado.insert(ttk.END, "No hay fallas únicas por día que cumplan los filtros.\n")

        else:
            text_resultado.insert(ttk.END, "No se encontraron fallas con los filtros dados.\n")


    def exportar():
        filtro_srv = combo_servidor.get().strip()
        filtro_job = combo_trabajo.get().strip()
        filtro_fecha = combo_fecha.get().strip()
        filtro_estado = combo_estado.get().strip()

        resultados_para_exportar = []
        # Volvemos a filtrar los 'backups_por_servidor' ya pre-filtrados por fallas
        for servidor, historial in backups_por_servidor.items():
            if filtro_srv and filtro_srv != servidor:
                continue
            for intento in historial:
                # Ya sabemos que son fallas, solo aplicamos los filtros de la UI
                if filtro_job and filtro_job != intento['nombre_trabajo']:
                    continue
                if filtro_fecha and filtro_fecha != intento['inicio'].strftime('%Y-%m-%d'):
                    continue
                if filtro_estado and filtro_estado != intento['estado']:
                    continue
                resultados_para_exportar.append({**intento, 'servidor': servidor})

        if resultados_para_exportar:
            # Exportar solo las fallas únicas por día para el informe
            fallas_distinct_para_exportar = {}
            for r in resultados_para_exportar:
                clave_diaria = (r['servidor'], r['nombre_trabajo'], r['inicio'].date())
                if clave_diaria not in fallas_distinct_para_exportar or r['inicio'] > fallas_distinct_para_exportar[clave_diaria]['inicio']:
                    fallas_distinct_para_exportar[clave_diaria] = r
            
            exportar_excel(list(fallas_distinct_para_exportar.values()))
        else:
            ttk.messagebox.showinfo("Exportar", "No hay fallas para exportar con los filtros dados.")

    app.mainloop()


# --- MAIN ---
if __name__ == "__main__":
    try:
        print("Cargando datos...")
        # Llama a analizar_informes al inicio para cargar todos los datos
        backups_servidor, trabajos_por_servidor, trabajos, servidores, fechas, fechas_por_servidor_y_trabajo = analizar_informes(RUTA_INFORMES)
        
        # Si no se cargan servidores, significa que no se encontraron datos válidos
        if not servidores:
            print("ADVERTENCIA: No se encontraron datos válidos en los informes. Asegúrate de que los archivos existan y contengan los encabezados correctos.")
            # Puedes añadir un mensaje emergente aquí si lo deseas:
            # ttk.messagebox.showwarning("Advertencia", "No se encontraron datos válidos en los informes. Asegúrate de que los archivos existan y contengan los encabezados correctos.")

        crear_interfaz(backups_servidor, trabajos_por_servidor, trabajos, servidores, fechas, fechas_por_servidor_y_trabajo)
    except Exception as e:
        logging.error(f"Error general en la aplicación: {e}")
        print(f"ERROR: Fallo en la aplicación: {e}")
